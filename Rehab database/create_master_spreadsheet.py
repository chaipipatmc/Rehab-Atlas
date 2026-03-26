import json
import os
from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.utils import get_column_letter

BASE = os.path.dirname(os.path.abspath(__file__))

# File -> Continent mapping
FILE_MAP = {
    "asia_centers_data.json": "Asia",
    "europe_centers_data.json": "Europe",
    "africa_centers_data.json": "Africa",
    "latam_centers_data.json": "Latin America",
    "canada_centers_data.json": "Canada",
    "australia_centers_data.json": "Australia",
}

# Field name normalization map
FIELD_MAP = {
    "centername": "centerName",
    "center_name": "centerName",
    "name": "centerName",
    "recoveryurl": "recoveryUrl",
    "recovery_url": "recoveryUrl",
    "url": "recoveryUrl",
}

def normalize_record(rec):
    normalized = {}
    for k, v in rec.items():
        key_lower = k.lower().replace(" ", "")
        mapped = FIELD_MAP.get(key_lower, k)
        normalized[mapped] = (v or "").strip() if isinstance(v, str) else (v or "")
    return normalized

# Load all data
all_records = []
for filename, continent in FILE_MAP.items():
    path = os.path.join(BASE, filename)
    with open(path, "r", encoding="utf-8") as f:
        data = json.load(f)
    for rec in data:
        nr = normalize_record(rec)
        nr["continent"] = continent
        all_records.append(nr)

# Sort by Continent, then Country/Region, then Center Name
all_records.sort(key=lambda r: (
    r.get("continent", ""),
    r.get("region", ""),
    r.get("centerName", "")
))

# Save combined JSON
with open(os.path.join(BASE, "master_all_centers.json"), "w", encoding="utf-8") as f:
    json.dump(all_records, f, ensure_ascii=False, indent=2)

# ---------- Create Excel ----------
wb = Workbook()

# Colors & styles
DARK_BLUE = "1F3864"
LIGHT_BLUE = "F2F7FB"
YELLOW_HIGHLIGHT = "FFF2CC"
GREEN_HIGHLIGHT = "E2EFDA"
LINK_COLOR = "0563C1"
WHITE = "FFFFFF"

header_font = Font(name="Arial", size=11, bold=True, color=WHITE)
header_fill = PatternFill(start_color=DARK_BLUE, end_color=DARK_BLUE, fill_type="solid")
data_font = Font(name="Arial", size=10)
link_font = Font(name="Arial", size=10, color=LINK_COLOR, underline="single")
alt_fill = PatternFill(start_color=LIGHT_BLUE, end_color=LIGHT_BLUE, fill_type="solid")
yellow_fill = PatternFill(start_color=YELLOW_HIGHLIGHT, end_color=YELLOW_HIGHLIGHT, fill_type="solid")
green_fill = PatternFill(start_color=GREEN_HIGHLIGHT, end_color=GREEN_HIGHLIGHT, fill_type="solid")
no_fill = PatternFill(fill_type=None)
thin_border = Border(
    left=Side(style="thin", color="D9D9D9"),
    right=Side(style="thin", color="D9D9D9"),
    top=Side(style="thin", color="D9D9D9"),
    bottom=Side(style="thin", color="D9D9D9"),
)

COLUMNS = ["#", "Continent", "Country/Region", "Center Name", "Website", "Phone", "Email", "Recovery.com URL"]
COL_WIDTHS = [5, 14, 18, 42, 45, 22, 35, 60]

# ===== Sheet 1: All Centers =====
ws = wb.active
ws.title = "All Centers"
ws.freeze_panes = "A2"

# Headers
for col_idx, (header, width) in enumerate(zip(COLUMNS, COL_WIDTHS), 1):
    cell = ws.cell(row=1, column=col_idx, value=header)
    cell.font = header_font
    cell.fill = header_fill
    cell.alignment = Alignment(horizontal="center", vertical="center")
    cell.border = thin_border
    ws.column_dimensions[get_column_letter(col_idx)].width = width

# Data rows
for row_idx, rec in enumerate(all_records, 2):
    seq = row_idx - 1
    website = rec.get("website", "")
    phone = rec.get("phone", "")
    email = rec.get("email", "")
    recovery_url = rec.get("recoveryUrl", "")

    values = [
        seq,
        rec.get("continent", ""),
        rec.get("region", ""),
        rec.get("centerName", ""),
        website,
        phone,
        email,
        recovery_url,
    ]

    is_alt = (row_idx % 2 == 0)
    row_fill = alt_fill if is_alt else no_fill

    for col_idx, val in enumerate(values, 1):
        cell = ws.cell(row=row_idx, column=col_idx, value=val)
        cell.font = data_font
        cell.fill = row_fill
        cell.border = thin_border
        cell.alignment = Alignment(vertical="center")

    # Website column (5) - yellow if missing, link font if present
    ws_cell = ws.cell(row=row_idx, column=5)
    if not website:
        ws_cell.fill = yellow_fill
    else:
        ws_cell.font = link_font

    # Email column (7) - green if present, link font if present
    em_cell = ws.cell(row=row_idx, column=7)
    if email:
        em_cell.fill = green_fill
        em_cell.font = link_font

    # Recovery URL column (8) - link font if present
    url_cell = ws.cell(row=row_idx, column=8)
    if recovery_url:
        url_cell.font = link_font

# Auto-filter
ws.auto_filter.ref = f"A1:H{len(all_records) + 1}"

# ===== Sheet 2: Summary =====
ws2 = wb.create_sheet("Summary")

# Compute stats per continent
continent_order = ["Africa", "Asia", "Australia", "Canada", "Europe", "Latin America"]
stats = {}
for c in continent_order:
    recs = [r for r in all_records if r.get("continent") == c]
    stats[c] = {
        "total": len(recs),
        "with_website": sum(1 for r in recs if r.get("website")),
        "with_email": sum(1 for r in recs if r.get("email")),
        "with_phone": sum(1 for r in recs if r.get("phone")),
    }

summary_headers = ["Continent", "Total Centers", "With Website", "With Email", "With Phone"]
summary_widths = [18, 16, 16, 16, 16]

ws2.freeze_panes = "A2"

for col_idx, (h, w) in enumerate(zip(summary_headers, summary_widths), 1):
    cell = ws2.cell(row=1, column=col_idx, value=h)
    cell.font = header_font
    cell.fill = header_fill
    cell.alignment = Alignment(horizontal="center", vertical="center")
    cell.border = thin_border
    ws2.column_dimensions[get_column_letter(col_idx)].width = w

for row_idx, continent in enumerate(continent_order, 2):
    s = stats[continent]
    vals = [continent, s["total"], s["with_website"], s["with_email"], s["with_phone"]]
    is_alt = (row_idx % 2 == 0)
    row_fill = alt_fill if is_alt else no_fill
    for col_idx, v in enumerate(vals, 1):
        cell = ws2.cell(row=row_idx, column=col_idx, value=v)
        cell.font = data_font
        cell.fill = row_fill
        cell.border = thin_border
        cell.alignment = Alignment(horizontal="center" if col_idx > 1 else "left", vertical="center")

# Totals row
total_row = len(continent_order) + 2
total_vals = [
    "TOTAL",
    sum(s["total"] for s in stats.values()),
    sum(s["with_website"] for s in stats.values()),
    sum(s["with_email"] for s in stats.values()),
    sum(s["with_phone"] for s in stats.values()),
]
for col_idx, v in enumerate(total_vals, 1):
    cell = ws2.cell(row=total_row, column=col_idx, value=v)
    cell.font = Font(name="Arial", size=10, bold=True)
    cell.fill = PatternFill(start_color="D6E4F0", end_color="D6E4F0", fill_type="solid")
    cell.border = thin_border
    cell.alignment = Alignment(horizontal="center" if col_idx > 1 else "left", vertical="center")

ws2.auto_filter.ref = f"A1:E{total_row}"

# Save
output_path = os.path.join(BASE, "Master Rehab Centers Database.xlsx")
wb.save(output_path)

# Print stats
print(f"\n{'='*60}")
print(f"  Master Rehab Centers Database - Created Successfully")
print(f"{'='*60}")
print(f"\n  Total records: {len(all_records)}")
print(f"\n  Per-continent breakdown:")
for c in continent_order:
    s = stats[c]
    print(f"    {c:16s}  {s['total']:4d} centers | {s['with_website']:3d} websites | {s['with_email']:3d} emails | {s['with_phone']:3d} phones")
print(f"    {'─'*72}")
print(f"    {'TOTAL':16s}  {total_vals[1]:4d} centers | {total_vals[2]:3d} websites | {total_vals[3]:3d} emails | {total_vals[4]:3d} phones")
print(f"\n  Output files:")
print(f"    - {output_path}")
print(f"    - {os.path.join(BASE, 'master_all_centers.json')}")
print(f"{'='*60}")
