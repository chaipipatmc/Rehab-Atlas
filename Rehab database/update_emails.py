import json
import openpyxl
from openpyxl.styles import Font
from urllib.parse import urlparse

email_map = {
    "algarve-wellness.org": "melissa@algarve-wellness.org",
    "beroendekliniken.se": "info@beroendekliniken.se",
    "blueroomrecovery.se": "info@blueroomrecovery.se",
    "bonairerecovery.com": "info@bonairerecovery.com",
    "brainm.eu": "info@brainm.eu",
    "castlecraig.se": "info@beroendekliniken.se",
    "dianova.pt": "secretariado@dianova.pt",
    "dozepassos.pt": "info@dozepassos.pt",
    "ggzinterventie.nl": "info@ggzinterventie.nl",
    "harborlondon.com": "talktous@harborlondon.com",
    "harmonyathens.com": "info@harmonyathens.com",
    "hellenicpractice.com": "info@hellenicpractice.com",
    "heritagecounseling.pt": "geral@heritagecounseling.pt",
    "horizonterecover.pt": "geral@horizonterecover.org",
    "innerliferecovery.com": "info@innerliferecovery.com",
    "khironclinics.com": "help@khironclinics.com",
    "kusnachtpractice.com": "info@kusnachtpractice.ch",
    "neoviva.com": "info@neoviva.com",
    "nhrc.ie": "info@nhrc.ie",
    "nightingalehospital.co.uk": "info@nightingalehospital.co.uk",
    "oasispremiumrecovery.com": "admissions@oasispremiumrecovery.com",
    "paracelsus-recovery.com": "info@paracelsus-recovery.com",
    "poseidonmethod.com": "info@poseidonmethod.com",
    "promis.co.uk": "enquiries@promisclinics.com",
    "recoverylighthouse.com": "help@recoverylighthouse.com",
    "resetpmd.com": "hello@reset.se",
    "rosglasrecovery.com": "info@rosglasrecovery.com",
    "step1recovery.com": "info@step1recovery.com",
    "stepstogether.co.uk": "pressandmarketing@stepstogether.co.uk",
    "thebalance.clinic": "spain@thebalance.clinic",
    "thebridgemarbella.com": "info@thebridgemarbella.com",
    "theretreatinitaly.com": "info@theretreatinitaly.com",
    "www.affect2u.be": "contact@affect2u.be",
    "www.castlecraig.co.uk": "info@castlecraig.co.uk",
    "www.cuanmhuire.ie": "donorrelations@cuanmhuire.ie",
    "www.denrooyclinics.com": "info@denrooyclinics.com",
    "www.ggzmomentum.nl": "ggzminfo@vigogroep.nl",
    "www.ibizacalm.com": "info@ibizacalm.com",
    "www.inharmonirehab.com": "info@inharmonirehab.com",
    "www.jugendhilfe.de": "info@jugendhilfe.de",
    "www.lametairie.ch": "info@lametairie.ch",
    "www.libertyhouseclinic.co.uk": "info@libertyhouseclinic.co.uk",
    "www.linwoodhouse.co.uk": "help@linwoodhouse.co.uk",
    "www.marievahealthcare.com": "info@marievahealthcare.com",
    "www.mywaybettyford.de": "info@mywaybettyford.de",
    "www.priorygroup.com": "info@priorygroup.com",
    "www.reset.be": "info@reset.be",
    "www.rutlandcentre.ie": "info@rutlandcentre.ie",
    "www.sanctuarylodge.com": "info@sanctuarylodge.com",
    "www.sanctuaryoceanic.com": "info@thesanctuaryworld.com",
    "www.sanctuaryunitedkingdom.com": "info@thesanctuaryworld.com",
    "www.smarmorecastle.ie": "info@smarmorecastle.ie",
    "www.smarttms.co.uk": "info@smarttms.co.uk",
    "www.tabularasaretreat.com": "info@tabularasaretreat.com",
    "www.u-center.eu": "comcenter@u-center.nl",
    "www.villaparadisospain.com": "info@villaparadisospain.com",
    "zeusrehab.com": "pomoc@zeusrehab.com",
    "newlife.rehab": "info@newlife.rehab",
    "www.caminorecovery.com": "info@caminorecovery.com",
    "www.caritasmalta.org": "info@caritasmalta.org",
    "www.centrosannicola.com": "info@centrosannicola.com",
}

def extract_domain(url):
    """Extract domain (host) from URL, keeping www. prefix if present."""
    if not url:
        return None
    parsed = urlparse(url)
    host = parsed.netloc or parsed.path.split('/')[0]
    host = host.rstrip('/')
    return host.lower() if host else None

# --- Update JSON ---
with open('europe_centers_data.json', 'r', encoding='utf-8') as f:
    centers = json.load(f)

json_matched = 0
json_missing = 0
for center in centers:
    website = center.get('website')
    domain = extract_domain(website)
    email = email_map.get(domain) if domain else None
    if email:
        center['email'] = email
        json_matched += 1
    else:
        json_missing += 1

with open('europe_centers_data.json', 'w', encoding='utf-8') as f:
    json.dump(centers, f, indent=2, ensure_ascii=False)

print(f"JSON: {json_matched} matched, {json_missing} missing (total {len(centers)})")

# --- Update Excel ---
wb = openpyxl.load_workbook('Europe Rehab Centers.xlsx')
ws = wb.active
font = Font(name='Arial', size=10)

excel_matched = 0
excel_missing = 0
for row in range(2, ws.max_row + 1):
    website_cell = ws.cell(row=row, column=4).value  # Column D = Website
    domain = extract_domain(website_cell)
    email = email_map.get(domain) if domain else None
    if email:
        cell = ws.cell(row=row, column=7)  # Column G = Email
        cell.value = email
        cell.font = font
        excel_matched += 1
    else:
        excel_missing += 1

wb.save('Europe Rehab Centers.xlsx')
print(f"Excel: {excel_matched} matched, {excel_missing} missing (total {ws.max_row - 1})")
print("Done! Both files updated.")
