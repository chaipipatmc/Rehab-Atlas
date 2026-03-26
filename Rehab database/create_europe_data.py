import json
import openpyxl
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.utils import get_column_letter

# All 147 centers from 4 batches
all_centers = [
# BATCH 1
{"recoveryUrl":"https://recovery.com/affect2u-vosselaar-belgium/","centerName":"Affect2U","website":"https://www.affect2u.be/en/","phone":"+44(1865)638477"},
{"recoveryUrl":"https://recovery.com/algarve-wellness-portugal/","centerName":"Algarve Wellness","website":"https://algarve-wellness.org/","phone":"+442045725855"},
{"recoveryUrl":"https://recovery.com/den-roy-clinics-meerle-belgium/","centerName":"Den Rooy Clinics","website":"https://www.denrooyclinics.com/nl/","phone":"+32(0)32937998"},
{"recoveryUrl":"https://recovery.com/athena-passages-greece/","centerName":"Athena Passages","website":"https://athenapassages.org/","phone":"+30(211)198-0645"},
{"recoveryUrl":"https://recovery.com/blaues-kreuz-cologne-germany/","centerName":"Blaues Kreuz Cologne","website":"https://www.blaues-kreuz.de/de/rheinland/koeln/blaukreuz-zentrum-koeln/","phone":"+49091638135"},
{"recoveryUrl":"https://recovery.com/blueroom-recovery-stockholm-sweden/","centerName":"Blueroom Recovery","website":"https://blueroomrecovery.se/","phone":"+468982860"},
{"recoveryUrl":"https://recovery.com/dianova-addiction-recovery-centre/","centerName":"Dianova Addiction Recovery Centre","website":"https://dianova.pt/en/discover-quinta-das-lapas/","phone":"+351911510098"},
{"recoveryUrl":"https://recovery.com/doze-passos-portugal/","centerName":"Doze Passos","website":"https://dozepassos.pt/","phone":"+310855001212"},
{"recoveryUrl":"https://recovery.com/bonaire-mallorca-spain/","centerName":"Bonaire Bespoke Recovery","website":"https://bonairerecovery.com/","phone":"+44(23)94330051"},
{"recoveryUrl":"https://recovery.com/drug-therapy-center-lago-berlin-germany/","centerName":"Drug Therapy Center Berlin LAGO","website":"https://www.berlin-suchthilfe.de","phone":"+49308068840"},
{"recoveryUrl":"https://recovery.com/calda-clinic/","centerName":"CALDA Clinic","website":"http://www.caldaclinic.com","phone":"+41442242000"},
{"recoveryUrl":"https://recovery.com/drug-therapy-center-transitional-housing-berlin-germany/","centerName":"Drug Therapy Center Berlin Transitional Housing","website":"https://www.berlin-suchthilfe.de/einrichtungen/uebergangseinrichtung/","phone":"+493029385200"},
{"recoveryUrl":"https://recovery.com/fachklinik-hamburg-germany/","centerName":"Fachklinik Hamburg-Mitte","website":"https://www.jugendhilfe.de/fachklinik/","phone":"+4940570025250"},
{"recoveryUrl":"https://recovery.com/ggz-interventie-amsterdam-netherlands/","centerName":"GGZ Interventie","website":"https://ggzinterventie.nl/","phone":"+31202310000"},
{"recoveryUrl":"https://recovery.com/calme-france/","centerName":"Calme France","website":"https://www.calme.fr/index.php","phone":"33(0)237916333"},
{"recoveryUrl":"https://recovery.com/camino-recovery/","centerName":"Camino Recovery","website":"https://www.caminorecovery.com/","phone":"+44(1428)774943"},
{"recoveryUrl":"https://recovery.com/ggz-momentum-breda-netherlands/","centerName":"GGZ Momentum","website":"https://www.ggzmomentum.nl/","phone":"0763040015"},
{"recoveryUrl":"https://recovery.com/brainm-clinics-antwerp/","centerName":"BRAINM Clinics Antwerp","website":"https://brainm.eu/","phone":"+3232830223"},
{"recoveryUrl":"https://recovery.com/caritas-malta-san-blas-zebbug-malta/","centerName":"Caritas Malta - San Blas Therapeutic Community","website":"https://www.caritasmalta.org/","phone":"+35622199600"},
{"recoveryUrl":"https://recovery.com/harbor-london/","centerName":"Harbor London","website":"https://harborlondon.com","phone":"+44(20)45389313"},
{"recoveryUrl":"https://recovery.com/castle-craig/","centerName":"Castle Craig","website":"https://www.castlecraig.co.uk/","phone":"+44(131)3810190"},
{"recoveryUrl":"https://recovery.com/catch-recovery-dublin-ireland/","centerName":"CATCH Recovery Dublin","website":"https://www.catchrecovery.ie/","phone":"+35312707603"},
{"recoveryUrl":"https://recovery.com/harmony-place-athens-greece/","centerName":"Harmony Place Athens","website":"https://harmonyathens.com/","phone":"+306932651751"},
{"recoveryUrl":"https://recovery.com/castle-craig-sweden/","centerName":"Castle Craig Sweden","website":"https://castlecraig.se/","phone":"46842002828"},
{"recoveryUrl":"https://recovery.com/changes-ggz-halsteren-netherlands/","centerName":"Changes GGZ Halsteren","website":"https://changesggz.nl/locaties/halsteren/","phone":"0882426437"},
{"recoveryUrl":"https://recovery.com/hellenic-practice/","centerName":"Hellenic Practice","website":"https://hellenicpractice.com","phone":"(833)377-2692"},
{"recoveryUrl":"https://recovery.com/changes-ggz-breda/","centerName":"Changes GGZ Breda","website":"https://changesggz.nl/locaties/breda/","phone":"0882426437"},
{"recoveryUrl":"https://recovery.com/heritage-counseling-portugal/","centerName":"Heritage Counseling Clinic","website":"https://heritagecounseling.pt","phone":"+351210145575"},
{"recoveryUrl":"https://recovery.com/changes-ggz-weert/","centerName":"Changes GGZ Weert","website":"https://changesggz.nl/locaties/weert/","phone":"0882426437"},
{"recoveryUrl":"https://recovery.com/clinique-la-metairie-switzerland/","centerName":"Clinique La Métairie","website":"https://www.lametairie.ch/","phone":"+41223632020"},
{"recoveryUrl":"https://recovery.com/cuan-mhuire-bruree/","centerName":"Cuan Mhuire Bruree","website":"https://www.cuanmhuire.ie","phone":"+353(0)6390555"},
{"recoveryUrl":"https://recovery.com/cuan-mhuire-athy/","centerName":"Cuan Mhuire - Athy","website":"https://www.cuanmhuire.ie","phone":"+353(0)598631493"},
{"recoveryUrl":"https://recovery.com/clinic-les-alpes/","centerName":"Clinic Les Alpes","website":"https://cliniclesalpes.com","phone":"+41215391922"},
{"recoveryUrl":"https://recovery.com/cuan-mhuire-coolarne-ireland/","centerName":"Cuan Mhuire Coolarne","website":"https://www.cuanmhuire.ie","phone":"353(0)91797102"},
{"recoveryUrl":"https://recovery.com/cuan-mhuire-farnanes-ireland/","centerName":"Cuan Mhuire Farnanes","website":"https://www.cuanmhuire.ie","phone":"+353(0)217335994"},
{"recoveryUrl":"https://recovery.com/delamere-england/","centerName":"Delamere Health","website":"https://delamere.com/","phone":"+441843264336"},
{"recoveryUrl":"https://recovery.com/cuan-mhuire-newry-ireland/","centerName":"Cuan Mhuire Newry","website":"https://www.cuanmhuire.ie","phone":"0044(0)2830849010"},
# BATCH 2
{"recoveryUrl":"https://recovery.com/outpatient/good-hope-clinic-netherlands-amsterdam-netherlands/","centerName":"Good Hope Clinic Netherlands","website":"","phone":"+310208543459"},
{"recoveryUrl":"https://recovery.com/outpatient/ready-for-change-rotterdam/","centerName":"Ready For Change","website":"","phone":"+31887323942"},
{"recoveryUrl":"https://recovery.com/outpatient/reconext-wittem-netherlands/","centerName":"Reconext","website":"","phone":"+31850776330"},
{"recoveryUrl":"https://recovery.com/outpatient/solutions-rotterdam-netherlands/","centerName":"SolutionS Rotterdam","website":"","phone":"033-2048550"},
{"recoveryUrl":"https://recovery.com/holina-youth-village-cyprus/","centerName":"Holina Village For Ages 16-25","website":"https://holinacyprus.com","phone":"+66626418369"},
{"recoveryUrl":"https://recovery.com/hopital-marmottan-paris-france/","centerName":"Hopital Marmottan","website":"https://www.hopital-marmottan.fr/wordpress/","phone":"0156687030"},
{"recoveryUrl":"https://recovery.com/outpatient/victory-village-drachten-netherlands/","centerName":"Victory Village","website":"","phone":"31858227745"},
{"recoveryUrl":"https://recovery.com/paracelsus-recovery/","centerName":"Paracelsus Recovery","website":"https://paracelsus-recovery.com","phone":"+41-44-505-11-68"},
{"recoveryUrl":"https://recovery.com/outpatient/spoor6-bussum-netherlands/","centerName":"Spoor6","website":"","phone":"035-6975400"},
{"recoveryUrl":"https://recovery.com/horizonte-sintra-portugal/","centerName":"Horizonte","website":"https://horizonterecover.pt/","phone":"(+351)210537970"},
{"recoveryUrl":"https://recovery.com/ibiza-calm/","centerName":"Ibiza Calm","website":"https://www.ibizacalm.com/","phone":"44(1793)540238"},
{"recoveryUrl":"https://recovery.com/paros-practice-paros-greece/","centerName":"Paros Practice","website":"https://parospractice.com","phone":"+306984453691"},
{"recoveryUrl":"https://recovery.com/parachute/","centerName":"Parachute Luxury Treatment","website":"https://www.parachutevienna.com","phone":"+432252254530"},
{"recoveryUrl":"https://recovery.com/inharmoni-marbella-spain/","centerName":"InHarmoni Rehab Marbella","website":"https://www.inharmonirehab.com/","phone":"+44(1789)643665"},
{"recoveryUrl":"https://recovery.com/priory-wellbeing-harley-street/","centerName":"Priory Wellbeing Centre Harley Street","website":"https://www.priorygroup.com/locations/priory-wellbeing-centre-harley-street","phone":"+442070790555"},
{"recoveryUrl":"https://recovery.com/innerlife-recovery-marbella-spain/","centerName":"InnerLife Recovery","website":"https://innerliferecovery.com/","phone":"+34613795340"},
{"recoveryUrl":"https://recovery.com/privatklinik-meiringen-switzerland/","centerName":"Privatklinik Meiringen","website":"https://www.privatklinik-meiringen.ch/","phone":"+41339728111"},
{"recoveryUrl":"https://recovery.com/khiron-clinics-oxfordshire-england-united-kingdom/","centerName":"Khiron Clinics","website":"https://khironclinics.com","phone":"+44(1403)337250"},
{"recoveryUrl":"https://recovery.com/promis-central-london/","centerName":"PROMIS London","website":"https://promis.co.uk/","phone":"+441223981457"},
{"recoveryUrl":"https://recovery.com/my-way-betty-ford-klinik/","centerName":"My Way Betty Ford Klinik","website":"https://www.mywaybettyford.de/","phone":"+498005574755"},
{"recoveryUrl":"https://recovery.com/neoviva-lake-lucerne-switzerland/","centerName":"NEOVIVA","website":"https://neoviva.com","phone":"+41-41-539-18-10"},
{"recoveryUrl":"https://recovery.com/marieva-healthcare/","centerName":"Marieva Healthcare","website":"https://www.marievahealthcare.com/","phone":"+302130249130"},
{"recoveryUrl":"https://recovery.com/new-hope-residential-centre-dublin-ireland/","centerName":"New Hope Residential Centre","website":"https://nhrc.ie/","phone":"+353(01)4512346"},
{"recoveryUrl":"https://recovery.com/new-life-prestige-program-marbella-spain/","centerName":"New Life Marbella Prestige Program","website":"https://newlife.rehab/location/worlds-luxury-rehab-marbella-club-prestige","phone":"+34601241130"},
{"recoveryUrl":"https://recovery.com/neoviva-womens-programme-switzerland/","centerName":"NEOVIVA Women's Programme","website":"https://neoviva.com/who-we-help/womens-programme/","phone":"+41-62-539-19-88"},
{"recoveryUrl":"https://recovery.com/new-life-marbella-spain/","centerName":"New Life Marbella","website":"https://newlife.rehab","phone":"+34600641248"},
{"recoveryUrl":"https://recovery.com/nightingale-hospital-london-united-kingdom/","centerName":"Nightingale Hospital","website":"https://nightingalehospital.co.uk","phone":"+448081754706"},
{"recoveryUrl":"https://recovery.com/oasis-premium-recovery/","centerName":"Oasis Premium Recovery","website":"https://oasispremiumrecovery.com","phone":"+44(1634)980403"},
{"recoveryUrl":"https://recovery.com/outpatient/brainm-clinics-ghent-south/","centerName":"BRAINM Clinics Ghent-South","website":"","phone":"+3232830223"},
{"recoveryUrl":"https://recovery.com/outpatient/castle-craig-rotterdam-netherlands/","centerName":"Castle Craig Rotterdam","website":"","phone":"+31887707077"},
{"recoveryUrl":"https://recovery.com/notre-abri/","centerName":"Notre Abri","website":"https://www.addiction-abstinence.org","phone":"0041(0)774220476"},
{"recoveryUrl":"https://recovery.com/outpatient/connection-sggz-meerlo-netherlands/","centerName":"Connection SGGZ","website":"","phone":"+31040-3035023"},
{"recoveryUrl":"https://recovery.com/outpatient/csapa-oppelia-trait-dunion-villeneuve-la-garenne-france/","centerName":"CSAPA - Oppelia Trait d'Union Villeneuve-La-Garenne","website":"","phone":"+33141217350"},
{"recoveryUrl":"https://recovery.com/outpatient/die-gesundheitsgreisslerei-vienna/","centerName":"Die Gesundheitsgreisslerei","website":"","phone":"+43(01)8904543"},
{"recoveryUrl":"https://recovery.com/outpatient/csapa-oppelia-trait-dunion-boulogne-billancourt-france/","centerName":"CSAPA - Oppelia Trait d'Union Boulogne-Billancourt","website":"","phone":"+33141419801"},
{"recoveryUrl":"https://recovery.com/outpatient/gaia-paris/","centerName":"Gaia Paris","website":"","phone":"0177722200"},
{"recoveryUrl":"https://recovery.com/outpatient/brainm-clinics-amsterdam/","centerName":"BRAINM Clinics Amsterdam","website":"","phone":"+3232830223"},
# BATCH 3
{"recoveryUrl":"https://recovery.com/residential/priory-group-the-residence/","centerName":"The Residence","website":"https://www.priorygroup.com/the-residence/therapy-programmes","phone":"03330601538"},
{"recoveryUrl":"https://recovery.com/residential/phoenix-programmes/","centerName":"Phoenix Programmes S.L.","website":"","phone":"+34691881422"},
{"recoveryUrl":"https://recovery.com/residential/salux-care-group-belgium/","centerName":"Salux Care Group","website":"","phone":""},
{"recoveryUrl":"https://recovery.com/promis-hay-farm-united-kingdom/","centerName":"PROMIS Hay Farm","website":"https://promis.co.uk/about/clinics/rehab-rehabilitation-clinic/","phone":"+44(808)1964872"},
{"recoveryUrl":"https://recovery.com/residential/schweizer-haus-hadersdorf-main-house/","centerName":"Schweizer Haus Hadersdorf - The Main House","website":"","phone":"+43(01)9791083"},
{"recoveryUrl":"https://recovery.com/reset-belgium/","centerName":"Reset Belgium","website":"https://www.reset.be/","phone":"033458070"},
{"recoveryUrl":"https://recovery.com/residential/schweizer-haus-hadersdorf-villa-atlantis/","centerName":"Schweizer Haus Hadersdorf - The Villa Atlantis","website":"","phone":"+43(01)9791083"},
{"recoveryUrl":"https://recovery.com/residential/the-edge-crete/","centerName":"The Edge Crete","website":"","phone":"+6620385469"},
{"recoveryUrl":"https://recovery.com/reset-pmd-djursholm-sweden/","centerName":"Reset Private Medical Detox","website":"https://resetpmd.com/","phone":"+46761646810"},
{"recoveryUrl":"https://recovery.com/residential/blaues-kreuz-coburg-germany/","centerName":"Blaues Kreuz Coburg","website":"","phone":"+49084219088133"},
{"recoveryUrl":"https://recovery.com/residential/blaues-kreuz-eichstatt-germany/","centerName":"Blaues Kreuz Eichstatt","website":"","phone":"+49084219088133"},
{"recoveryUrl":"https://recovery.com/residential/therapiestation-erlenhof-taubing-austria/","centerName":"Therapiestation Erlenhof","website":"","phone":"+43072776913"},
{"recoveryUrl":"https://recovery.com/residential/ukat-banbury-lodge-banbury-united-kingdom/","centerName":"Banbury Lodge","website":"","phone":"+44(24)75420678"},
{"recoveryUrl":"https://recovery.com/residential/blaues-kreuz-hagen-germany/","centerName":"Blaues Kreuz Hagen","website":"","phone":"+49023314827-0"},
{"recoveryUrl":"https://recovery.com/residential/ukat-primrose-lodge-guildford-united-kingdom/","centerName":"Primrose Lodge","website":"","phone":"+44(141)6738576"},
{"recoveryUrl":"https://recovery.com/residential/blaues-kreuz-lippe-germany/","centerName":"Blaues Kreuz Lippe","website":"","phone":"+49034646156-373"},
{"recoveryUrl":"https://recovery.com/rosglas-recovery/","centerName":"RosGlas Recovery","website":"https://rosglasrecovery.com/","phone":"+44(118)2303352"},
{"recoveryUrl":"https://recovery.com/residential/blaues-kreuz-mittweida-germany/","centerName":"Blaues Kreuz Mittweida","website":"","phone":"+4903727930579"},
{"recoveryUrl":"https://recovery.com/residential/viver-mais-braga-portugal/","centerName":"Viver Mais","website":"","phone":"+351253274940"},
{"recoveryUrl":"https://recovery.com/residential/blaues-kreuz-radevormwald-germany/","centerName":"Blaues Kreuz Curt-von-Knobelsdorff-Haus","website":"","phone":"+4902195672-0"},
{"recoveryUrl":"https://recovery.com/san-nicola-arcevia-italy/","centerName":"Centro San Nicola","website":"https://www.centrosannicola.com/en/home-2/","phone":"+3907319142"},
{"recoveryUrl":"https://recovery.com/residential/blaues-kreuz-sangerhausen-germany/","centerName":"Blaues Kreuz Sangerhausen","website":"","phone":"+49034646156-373"},
{"recoveryUrl":"https://recovery.com/residential/blaues-kreuz-stuttgart-germany/","centerName":"Blaues Kreuz Stuttgart","website":"","phone":"+4907112238088"},
{"recoveryUrl":"https://recovery.com/residential/blaues-kreuz-schindelbach-germany/","centerName":"Blaues Kreuz Schindelbach","website":"","phone":"+4903735939-0"},
{"recoveryUrl":"https://recovery.com/residential/breathe-life-recovery/","centerName":"Breathe Life Recovery","website":"","phone":""},
{"recoveryUrl":"https://recovery.com/residential/blaues-kreuz-wuppertal-germany/","centerName":"Blaues Kreuz Wuppertal","website":"","phone":"+49020262003-0"},
{"recoveryUrl":"https://recovery.com/residential/carlisle-house-belfast-ireland/","centerName":"Carlisle House","website":"","phone":"02890328308"},
{"recoveryUrl":"https://recovery.com/residential/caritas-malta-et-iris-zebbug-malta/","centerName":"Caritas Malta - Et Iris Female Therapeutic Community","website":"","phone":"+35622199600"},
{"recoveryUrl":"https://recovery.com/residential/fachklinik-eusserthal/","centerName":"Fachklinik Eusserthal","website":"","phone":"+490634520-0"},
{"recoveryUrl":"https://recovery.com/residential/blaues-kreuz-rauschenberg-germany/","centerName":"Blaues Kreuz Rauschenberg","website":"","phone":"+49091638135"},
{"recoveryUrl":"https://recovery.com/residential/foco-saude/","centerName":"Foco Saude","website":"","phone":"+44(141)6732735"},
{"recoveryUrl":"https://recovery.com/residential/go-green-project-x/","centerName":"Go Green Project X","website":"","phone":"+302111985985"},
{"recoveryUrl":"https://recovery.com/residential/istana-jiwa-ibiza/","centerName":"Istana Bespoke Health - Ibiza","website":"","phone":"+44(808)1688156"},
{"recoveryUrl":"https://recovery.com/residential/gylleby-treatment-home-sunne-sweden/","centerName":"Gylleby Treatment Home","website":"","phone":"+4656510230"},
{"recoveryUrl":"https://recovery.com/residential/momento-centro-de-tratamiento-de-adicciones-madrid-spain/","centerName":"MOMENTO Centro de Tratamiento de Adicciones","website":"","phone":"+34910801255"},
{"recoveryUrl":"https://recovery.com/residential/johannesbad-fachklinik-furth-im-wald-germany/","centerName":"Johannesbad Fachklinik Furth im Wald","website":"","phone":"+4909975020"},
{"recoveryUrl":"https://recovery.com/residential/narconon-argo-italy/","centerName":"Narconon Argo","website":"","phone":"800931151"},
# BATCH 4
{"recoveryUrl":"https://recovery.com/u-center-epen-netherlands/","centerName":"U Center","website":"https://www.u-center.eu/","phone":"0800-2224446"},
{"recoveryUrl":"https://recovery.com/ukat-liberty-house-luton-united-kingdom/","centerName":"Liberty House","website":"https://www.libertyhouseclinic.co.uk/","phone":"+44(1527)741775"},
{"recoveryUrl":"https://recovery.com/ukat-linwood-house-barnsley-united-kingdom/","centerName":"Linwood House","website":"https://www.linwoodhouse.co.uk/","phone":"+44(1254)457461"},
{"recoveryUrl":"https://recovery.com/ukat-oasis-bradford-bradford-united-kingdom/","centerName":"Oasis Recovery Bradford","website":"https://www.oasisrecovery.org.uk/","phone":"+44(1245)823991"},
{"recoveryUrl":"https://recovery.com/ukat-oasis-runcorn-runcorn-united-kingdom/","centerName":"Oasis Recovery Runcorn","website":"https://www.oasisrehab.co.uk/","phone":"+44(191)7431529"},
{"recoveryUrl":"https://recovery.com/satori-recovery-malaga-spain/","centerName":"Satori Recovery","website":"https://www.satorirecovery.com","phone":"+44(121)7282351"},
{"recoveryUrl":"https://recovery.com/ukat-recovery-lighthouse-worthing-united-kingdom/","centerName":"Recovery Lighthouse","website":"https://recoverylighthouse.com","phone":"+44(1823)218147"},
{"recoveryUrl":"https://recovery.com/sea-recovery-sotogrande-spain/","centerName":"Sea Recovery","website":"https://searecoverycentre.com/about-sea-recovery/confidentiality-commitment/","phone":"+44(1502)442907"},
{"recoveryUrl":"https://recovery.com/smarmore-castle/","centerName":"Smarmore Castle Private Clinic","website":"https://www.smarmorecastle.ie","phone":"+353-41-214-0627"},
{"recoveryUrl":"https://recovery.com/smart-tms-dublin-ireland/","centerName":"Smart TMS Dublin","website":"https://www.smarttms.co.uk/clinics/dublin/","phone":"353(01)2542514"},
{"recoveryUrl":"https://recovery.com/ukat-sanctuary-lodge-halstead-united-kingdom/","centerName":"Sanctuary Lodge","website":"https://www.sanctuarylodge.com/","phone":"+44(121)8161087"},
{"recoveryUrl":"https://recovery.com/vida-innova-portugal/","centerName":"Vida Innova","website":"https://vidainnova.com/","phone":"+351262143485"},
{"recoveryUrl":"https://recovery.com/step-one-recovery/","centerName":"Step One Recovery","website":"https://step1recovery.com","phone":"+447914760631"},
{"recoveryUrl":"https://recovery.com/villa-paradiso-rehab/","centerName":"Villa Paradiso Rehab Spain","website":"https://www.villaparadisospain.com/","phone":"ES:+34689806769"},
{"recoveryUrl":"https://recovery.com/steps-together-rainford-hall-united-kingdom/","centerName":"Steps Together - Rainford Hall","website":"https://stepstogether.co.uk/locations/rainford-hall/","phone":"+44(1325)521541"},
{"recoveryUrl":"https://recovery.com/stockholms-beroendeklinik-sweden/","centerName":"Stockholms Beroendeklinik","website":"https://beroendekliniken.se/","phone":"08-12400284"},
{"recoveryUrl":"https://recovery.com/villaramadas-portugal/","centerName":"VillaRamadas International Treatment Centre","website":"https://www.villaramadas.com/","phone":"(+351)918120945"},
{"recoveryUrl":"https://recovery.com/tabula-rasa-retreat/","centerName":"Tabula Rasa Retreat","website":"https://www.tabularasaretreat.com","phone":"+351965751649"},
{"recoveryUrl":"https://recovery.com/zeus-rehab-warsaw-poland/","centerName":"Zeus Detox Rehab & SPA","website":"https://zeusrehab.com/en/","phone":"+448081751986"},
{"recoveryUrl":"https://recovery.com/the-bridge-spain/","centerName":"The Bridge Marbella","website":"https://thebridgemarbella.com","phone":"+441522458190"},
{"recoveryUrl":"https://recovery.com/tal-ibwar-zebbug-malta/","centerName":"Tal-Ibwar (Caritas Malta)","website":"https://www.caritasmalta.org/services/tal-ibwar/about-tal-ibwar/","phone":"+35622199500"},
{"recoveryUrl":"https://recovery.com/the-house-stockholm-sweden/","centerName":"The House Stockholm","website":"https://thehouserehab.com/the-house-skeppsbron/","phone":"+46(0)76-6008030"},
{"recoveryUrl":"https://recovery.com/the-cottage/","centerName":"Life Works - The Cottage","website":"","phone":"01483754066"},
{"recoveryUrl":"https://recovery.com/the-retreat-in-italy/","centerName":"The Retreat in Italy","website":"https://theretreatinitaly.com/","phone":"+393333230381"},
{"recoveryUrl":"https://recovery.com/the-kusnacht-practice/","centerName":"Kusnacht Practice","website":"https://kusnachtpractice.com/","phone":"+41-44-505-64-76"},
{"recoveryUrl":"https://recovery.com/the-poseidon-method-greece/","centerName":"The Poseidon Method","website":"https://poseidonmethod.com/","phone":"+30(211)198-0645"},
{"recoveryUrl":"https://recovery.com/the-priory-hospital-bristol/","centerName":"Priory Hospital Bristol","website":"","phone":"+44(1223)650277"},
{"recoveryUrl":"https://recovery.com/the-rutland-centre-dublin-ireland/","centerName":"Rutland Centre","website":"https://www.rutlandcentre.ie","phone":"+35314946358"},
{"recoveryUrl":"https://recovery.com/the-sanctuary-world/","centerName":"The Sanctuary Oceanic","website":"https://www.sanctuaryoceanic.com/","phone":"(855)641-5431"},
{"recoveryUrl":"https://recovery.com/the-sanctuary-world-united-kingdom/","centerName":"The Sanctuary United Kingdom","website":"https://www.sanctuaryunitedkingdom.com/","phone":"+44(115)6472765"},
{"recoveryUrl":"https://recovery.com/the-youturn-netherlands/","centerName":"THE YOUTURN","website":"https://theyouturn.nl/","phone":"+31619188281"},
{"recoveryUrl":"https://recovery.com/thebalance-mallorca/","centerName":"THE BALANCE Rehab Clinic Mallorca","website":"https://thebalance.clinic/clinic-locations/mallorca/","phone":"+34871249003"},
{"recoveryUrl":"https://recovery.com/thebalance-marbella/","centerName":"","website":"","phone":""},
{"recoveryUrl":"https://recovery.com/thebalance-london/","centerName":"THE BALANCE Rehab Clinic - London","website":"https://thebalance.clinic/clinic-locations/london/","phone":"+442039961507"},
{"recoveryUrl":"https://recovery.com/thebalance-zurich/","centerName":"THE BALANCE Rehab Clinic - Zurich","website":"https://thebalance.clinic/de/","phone":"+41445005111"},
{"recoveryUrl":"https://recovery.com/triora-the-hague-netherlands/","centerName":"Triora","website":"https://triora.nl/","phone":"+31883583741"},
]

# Region assignment
def get_region(url):
    url_lower = url.lower()
    checks = [
        (["united-kingdom", "england", "london", "bristol", "scotland"], "United Kingdom"),
        (["spain", "marbella", "mallorca", "ibiza", "sotogrande", "malaga", "madrid"], "Spain"),
        (["switzerland", "zurich", "lucerne", "meiringen"], "Switzerland"),
        (["greece", "athens", "crete", "paros"], "Greece"),
        (["portugal", "sintra", "braga", "algarve"], "Portugal"),
        (["netherlands", "rotterdam", "amsterdam", "epen", "bussum", "drachten", "breda", "halsteren", "weert", "meerlo", "wittem"], "Netherlands"),
        (["ireland", "dublin", "belfast", "athy", "bruree", "coolarne", "farnanes", "newry"], "Ireland"),
        (["austria", "vienna", "taubing", "hadersdorf"], "Austria"),
        (["cyprus"], "Cyprus"),
        (["france", "paris", "calme"], "France"),
        (["italy", "arcevia"], "Italy"),
        (["germany", "berlin", "hamburg", "cologne", "coburg", "eichstatt", "hagen", "lippe", "mittweida", "radevormwald", "rauschenberg", "sangerhausen", "schindelbach", "stuttgart", "wuppertal", "eusserthal", "furth"], "Germany"),
        (["belgium", "vosselaar", "meerle", "antwerp", "ghent"], "Belgium"),
        (["sweden", "stockholm", "djursholm", "sunne"], "Sweden"),
        (["malta", "zebbug"], "Malta"),
        (["poland", "warsaw"], "Poland"),
    ]
    for keywords, region in checks:
        for kw in keywords:
            if kw in url_lower:
                return region
    return "Europe (Other)"

# Add region to each center
for c in all_centers:
    c["region"] = get_region(c["recoveryUrl"])

# Save combined JSON
with open("europe_centers_data.json", "w", encoding="utf-8") as f:
    json.dump(all_centers, f, indent=2, ensure_ascii=False)

print(f"Saved {len(all_centers)} centers to europe_centers_data.json")

# Sort by region then center name
all_centers.sort(key=lambda x: (x["region"].lower(), x["centerName"].lower()))

# Create Excel
wb = openpyxl.Workbook()
ws = wb.active
ws.title = "Europe Rehab Centers"

# Column widths (matching Asia file)
col_widths = {"A": 5, "B": 12, "C": 45, "D": 50, "E": 22, "F": 65, "G": 35}
for col, w in col_widths.items():
    ws.column_dimensions[col].width = w

# Headers
headers = ["#", "Region", "Center Name", "Website", "Phone", "Recovery.com URL", "Email"]
header_font = Font(bold=True, color="FFFFFF")
header_fill = PatternFill(start_color="2F5496", end_color="2F5496", fill_type="solid")

for col_idx, header in enumerate(headers, 1):
    cell = ws.cell(row=1, column=col_idx, value=header)
    cell.font = header_font
    cell.fill = header_fill
    cell.alignment = Alignment(horizontal="left", vertical="center")

# Alternating row colors
alt_fill = PatternFill(start_color="F2F7FB", end_color="F2F7FB", fill_type="solid")
yellow_fill = PatternFill(start_color="FFF2CC", end_color="FFF2CC", fill_type="solid")

for idx, center in enumerate(all_centers):
    row = idx + 2
    is_alt = (idx % 2 == 1)  # odd rows get alternating color (0-based)

    ws.cell(row=row, column=1, value=idx + 1)  # #
    ws.cell(row=row, column=2, value=center["region"])
    ws.cell(row=row, column=3, value=center["centerName"])
    ws.cell(row=row, column=4, value=center["website"])
    ws.cell(row=row, column=5, value=center["phone"])
    ws.cell(row=row, column=6, value=center["recoveryUrl"])
    ws.cell(row=row, column=7, value="")  # Email blank

    # Apply alternating fill to all cells in row
    for col_idx in range(1, 8):
        cell = ws.cell(row=row, column=col_idx)
        if is_alt:
            cell.fill = alt_fill

    # Yellow highlight for missing website (overrides alt fill)
    if not center["website"]:
        ws.cell(row=row, column=4).fill = yellow_fill

# Auto filter
ws.auto_filter.ref = f"A1:G{len(all_centers) + 1}"

# Freeze top row
ws.freeze_panes = "A2"

wb.save("Europe Rehab Centers.xlsx")
print(f"Saved Excel with {len(all_centers)} rows to 'Europe Rehab Centers.xlsx'")

# Print region summary
from collections import Counter
regions = Counter(c["region"] for c in all_centers)
print("\nRegion breakdown:")
for r, count in sorted(regions.items()):
    print(f"  {r}: {count}")
