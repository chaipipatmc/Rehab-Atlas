import json
import openpyxl
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.utils import get_column_letter
from urllib.parse import urlparse

# All 40 centers
centers = [
    {"recoveryUrl":"https://recovery.com/bellevue-sober-residence-cape-town/","centerName":"Bellevue Sober Residence","website":"https://bellevuesoberresidence.com/","phone":"0027(0)726500276"},
    {"recoveryUrl":"https://recovery.com/anker-huis-cape-town-south-africa/","centerName":"Anker Huis","website":"https://www.ankerhuis.co.za","phone":"+442045870495"},
    {"recoveryUrl":"https://recovery.com/cedars-midlands-estate-south-africa/","centerName":"The Cedars Midlands Estate","website":"https://thecedars.co.za/facility/cedars-midlands-estate/","phone":"+270818609578"},
    {"recoveryUrl":"https://recovery.com/changes-addiction-rehab-northcliff-south-africa/","centerName":"Changes Addiction Rehab","website":"https://changesrehab.co.za/","phone":"081-444-7000"},
    {"recoveryUrl":"https://recovery.com/connection-mental-healthcare-cape-town-south-africa/","centerName":"Connection Mental Healthcare","website":"https://connection-mhc.co.za/","phone":"+44(808)1967252"},
    {"recoveryUrl":"https://recovery.com/cherrywood-house-cape-town-south-africa/","centerName":"Cherrywood House","website":"https://www.cherrywoodhouse.com","phone":"(+31)208083758"},
    {"recoveryUrl":"https://recovery.com/freeman-house-north-west-south-africa/","centerName":"Freeman House Recovery","website":"https://www.freemanhouserecovery.com","phone":"+44(1727)222030"},
    {"recoveryUrl":"https://recovery.com/harmony-addictions-clinic/","centerName":"Harmony Clinic","website":"https://harmonyclinic.co.za/","phone":"+44(1895)546477"},
    {"recoveryUrl":"https://recovery.com/healing-wings-south-africa/","centerName":"Healing Wings","website":"https://www.healingwings.co.za/","phone":"0769603273"},
    {"recoveryUrl":"https://recovery.com/houghton-house-randburg/","centerName":"Houghton House Treatment Centres","website":"https://www.houghtonhouse.co.za/","phone":"+27-10-288-0859"},
    {"recoveryUrl":"https://recovery.com/eagles-view-little-falls-south-africa/","centerName":"Eagles View Wellness Centre","website":"https://evwc.co.za/eagles-view-programmes/halfway-programme/","phone":"+44(1224)007325"},
    {"recoveryUrl":"https://recovery.com/journey-recovery-wellness-south-africa/","centerName":"Journey Recovery & Wellness Centre","website":"https://journeyrecoverycentre.com/","phone":"+27824476727"},
    {"recoveryUrl":"https://recovery.com/les-mariannes-wellness-sanctuary/","centerName":"Les Mariannes Wellness Sanctuary","website":"https://lesmariannes.com/","phone":"+23057287110"},
    {"recoveryUrl":"https://recovery.com/myrehab-benoni-south-africa/","centerName":"MyRehab","website":"https://myrehab.co.za/","phone":"+27828863996"},
    {"recoveryUrl":"https://recovery.com/new-life-cape-town/","centerName":"New Life Cape Town","website":"https://newlife.rehab/location/luxury-rehab-capetown","phone":"(+27)212008424"},
    {"recoveryUrl":"https://recovery.com/ocean-bay-eastern-cape-south-africa/","centerName":"Ocean Bay Recovery","website":"https://www.oceanbayrecovery.com/","phone":"+27213002414"},
    {"recoveryUrl":"https://recovery.com/oasis-counseling-centre/","centerName":"Oasis Recovery Centre","website":"https://www.oasisrecoverycentre.com","phone":"+27(0)445331752"},
    {"recoveryUrl":"https://recovery.com/journey-ballito-south-africa/","centerName":"Journey Recovery & Wellness Ballito","website":"https://journeyballito.co.za","phone":"+44(808)1750760"},
    {"recoveryUrl":"https://recovery.com/liberty-home-cape-town-south-africa/","centerName":"Liberty Home - Weltevreden Manor","website":"https://libertyhomerehab.com","phone":"+27-87-250-2159"},
    {"recoveryUrl":"https://recovery.com/olive-hill-clinic-marrakech-morocco/","centerName":"Olive Hill Clinique","website":"https://olivehillclinic.com/","phone":"+212-6-61-56-62-75"},
    {"recoveryUrl":"https://recovery.com/on-course-recovery-johannesburg-south-africa/","centerName":"On Course Recovery and Wellness Centre","website":"https://oncourserecovery.co.za/","phone":"0827447747"},
    {"recoveryUrl":"https://recovery.com/reset-sober-living-south-africa/","centerName":"Reset Sober Living","website":"https://resetsa.co.za/index.html","phone":"+27655764489"},
    {"recoveryUrl":"https://recovery.com/residential/africa-wellness/","centerName":"Africa Wellness","website":"","phone":"+27847714566"},
    {"recoveryUrl":"https://recovery.com/residential/harmony-psychiatric-clinic/","centerName":"Harmony Psychiatric Clinic","website":"","phone":"+27217907779"},
    {"recoveryUrl":"https://recovery.com/residential/renewed-life-center-cape-town/","centerName":"Renewed Life Center","website":"","phone":"+27798800045"},
    {"recoveryUrl":"https://recovery.com/recovery-center-white-river-africa/","centerName":"Recovery Centre at White River","website":"https://whiteriverrecovery.com","phone":"+27-87-250-2751"},
    {"recoveryUrl":"https://recovery.com/residential/the-grange-treatment-center-youth-clinic-cape-town-south-africa/","centerName":"The Grange Treatment Centre - Youth Clinic","website":"","phone":"+27646690676"},
    {"recoveryUrl":"https://recovery.com/residential/the-grange-treatment-centre-adult-clinic-cape-town-south-africa/","centerName":"The Grange Treatment Centre","website":"","phone":"27-21-818-6185"},
    {"recoveryUrl":"https://recovery.com/residential/revivocare-cape-town-south-africa/","centerName":"Revivocare","website":"","phone":"+27794832728"},
    {"recoveryUrl":"https://recovery.com/rustenburg-addiction-care-south-africa/","centerName":"Rustenburg Addiction Care","website":"https://rustenburgaddictioncare.co.za/","phone":"+27(0)879432293"},
    {"recoveryUrl":"https://recovery.com/sandhurst-manor-cape-town-south-africa/","centerName":"Sandhurst Manor - Cape Town","website":"https://cpt.sandhurstmanor.com/","phone":"(+27)645520794"},
    {"recoveryUrl":"https://recovery.com/residential/villa-consano-cape-town/","centerName":"Villa Consano","website":"","phone":"(31)20-808-5484"},
    {"recoveryUrl":"https://recovery.com/sandhurst-manor-gauteng-south-africa/","centerName":"Sandhurst Manor - Johannesburg","website":"https://jhb.sandhurstmanor.com/","phone":"+27(0)615249652"},
    {"recoveryUrl":"https://recovery.com/stepping-stones-rehab-centre/","centerName":"Stepping Stones Rehab Centre","website":"https://akeso.co.za/clinic/akeso-stepping-stones","phone":"+27(0)217834230"},
    {"recoveryUrl":"https://recovery.com/the-cedars-cape-town-south-africa/","centerName":"The Cedars Cape Manor House","website":"https://thecedars.co.za/","phone":"+27-0818609578"},
    {"recoveryUrl":"https://recovery.com/south-coast-recovery-south-africa/","centerName":"South Coast Recovery Centre and Halfway House","website":"https://www.southcoastrecoverycentre.co.za/","phone":"+27393144777"},
    {"recoveryUrl":"https://recovery.com/villa-paradiso-tunisia/","centerName":"Villa Paradiso Tunisia","website":"https://www.villaparadisotunisia.com/","phone":"+21650477777"},
    {"recoveryUrl":"https://recovery.com/the-foundation-clinic-south-africa/","centerName":"The Foundation Clinic","website":"https://thefoundationclinic.com","phone":"+27(0)109003131"},
    {"recoveryUrl":"https://recovery.com/the-living-house-south-africa/","centerName":"The Living House","website":"https://thelivinghouse.co.za/","phone":"0844548464"},
    {"recoveryUrl":"https://recovery.com/white-river-manor/","centerName":"White River Manor","website":"","phone":"+27-87-250-2135"},
]

email_map = {
    "bellevuesoberresidence.com": "info@bellevuesoberresidence.com",
    "www.ankerhuis.co.za": "info@ankerhuis.co.za",
    "thecedars.co.za": "bookings@thecedars.co.za",
    "changesrehab.co.za": "reception@changesrehab.co.za",
    "connection-mhc.co.za": "info@connection-mhc.co.za",
    "www.cherrywoodhouse.com": "info@cherrywoodhouse.com",
    "www.freemanhouserecovery.com": "info@freemanhouserecovery.com",
    "harmonyclinic.co.za": "admissions@harmonyclinic.co.za",
    "www.healingwings.co.za": "help@healingwings.co.za",
    "www.houghtonhouse.co.za": "info@houghtonhouse.co.za",
    "evwc.co.za": "info@evwc.co.za",
    "journeyrecoverycentre.com": "admin@journeycentre.co.za",
    "myrehab.co.za": "admissions@myrehab.co.za",
    "www.oceanbayrecovery.com": "info@oceanbayrecovery.com",
    "www.oasisrecoverycentre.com": "info@oasiscentre.co.za",
    "journeyballito.co.za": "admin@journeycentre.co.za",
    "libertyhomerehab.com": "info@libertyhomerehab.com",
    "oncourserecovery.co.za": "",
    "whiteriverrecovery.com": "booking@whiteriverrecovery.com",
    "rustenburgaddictioncare.co.za": "enquiry@rustenburgcare.co.za",
    "cpt.sandhurstmanor.com": "info@sandhurstmanor.com",
    "jhb.sandhurstmanor.com": "info@sandhurstmanor.com",
    "www.southcoastrecoverycentre.co.za": "enquiries@scrc.co.za",
    "thefoundationclinic.com": "info@thefoundationclinic.co.za",
    "thelivinghouse.co.za": "info@thelivinghouse.co.za",
    "newlife.rehab": "info@newlife.rehab",
}

def get_region(url):
    url_lower = url.lower()
    if "morocco" in url_lower or "marrakech" in url_lower:
        return "Morocco"
    if "tunisia" in url_lower:
        return "Tunisia"
    if "les-mariannes" in url_lower or "mauritius" in url_lower:
        return "Mauritius"
    return "South Africa"

def get_email(website):
    if not website:
        return ""
    parsed = urlparse(website)
    domain = parsed.netloc or parsed.path
    domain = domain.rstrip("/")
    # Try exact match
    if domain in email_map:
        return email_map[domain]
    # Try without www
    if domain.startswith("www."):
        bare = domain[4:]
        if bare in email_map:
            return email_map[bare]
    # Try with www
    www = "www." + domain
    if www in email_map:
        return email_map[www]
    # Try base domain (for subdomains like cpt.sandhurstmanor.com)
    parts = domain.split(".")
    if len(parts) > 2:
        base = ".".join(parts[-2:])
        if base in email_map:
            return email_map[base]
    return ""

# Build enriched data
for c in centers:
    c["region"] = get_region(c["recoveryUrl"])
    c["email"] = get_email(c["website"])

# Sort by region then center name
centers.sort(key=lambda x: (x["region"], x["centerName"].lower()))

# Save JSON
with open("africa_centers_data.json", "w", encoding="utf-8") as f:
    json.dump(centers, f, indent=2, ensure_ascii=False)
print(f"Saved africa_centers_data.json with {len(centers)} centers")

# Create Excel
wb = openpyxl.Workbook()
ws = wb.active
ws.title = "Africa Rehab Centers"

headers = ["#", "Region", "Center Name", "Website", "Phone", "Recovery.com URL", "Email"]
header_font = Font(name="Calibri", bold=True, color="FFFFFF", size=11)
header_fill = PatternFill(start_color="2F5496", end_color="2F5496", fill_type="solid")
header_alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)

thin_border = Border(
    left=Side(style="thin", color="B4C6E7"),
    right=Side(style="thin", color="B4C6E7"),
    top=Side(style="thin", color="B4C6E7"),
    bottom=Side(style="thin", color="B4C6E7"),
)

# Write headers
for col_idx, header in enumerate(headers, 1):
    cell = ws.cell(row=1, column=col_idx, value=header)
    cell.font = header_font
    cell.fill = header_fill
    cell.alignment = header_alignment
    cell.border = thin_border

# Alternating row fills
light_fill = PatternFill(start_color="D6E4F0", end_color="D6E4F0", fill_type="solid")
white_fill = PatternFill(start_color="FFFFFF", end_color="FFFFFF", fill_type="solid")
yellow_fill = PatternFill(start_color="FFFF00", end_color="FFFF00", fill_type="solid")

# Write data
for idx, c in enumerate(centers):
    row = idx + 2
    row_fill = light_fill if idx % 2 == 0 else white_fill

    ws.cell(row=row, column=1, value=idx + 1).font = Font(name="Calibri", size=11)
    ws.cell(row=row, column=1).alignment = Alignment(horizontal="center")
    ws.cell(row=row, column=2, value=c["region"]).font = Font(name="Calibri", size=11)
    ws.cell(row=row, column=3, value=c["centerName"]).font = Font(name="Calibri", size=11)

    # Website - yellow if missing
    web_cell = ws.cell(row=row, column=4, value=c["website"])
    web_cell.font = Font(name="Calibri", size=11)
    if not c["website"]:
        web_cell.fill = yellow_fill

    ws.cell(row=row, column=5, value=c["phone"]).font = Font(name="Calibri", size=11)
    ws.cell(row=row, column=6, value=c["recoveryUrl"]).font = Font(name="Calibri", size=11)
    ws.cell(row=row, column=7, value=c["email"]).font = Font(name="Calibri", size=11)

    # Apply row fill and border
    for col in range(1, 8):
        cell = ws.cell(row=row, column=col)
        if col != 4 or c["website"]:  # Don't override yellow fill
            cell.fill = row_fill
        elif not c["website"]:
            cell.fill = yellow_fill
        cell.border = thin_border

# Column widths
col_widths = {"A": 5, "B": 14, "C": 42, "D": 55, "E": 22, "F": 65, "G": 32}
for col_letter, width in col_widths.items():
    ws.column_dimensions[col_letter].width = width

# Auto-filter
ws.auto_filter.ref = f"A1:G{len(centers) + 1}"

# Freeze panes
ws.freeze_panes = "A2"

wb.save("Africa Rehab Centers.xlsx")
print(f"Saved Africa Rehab Centers.xlsx with {len(centers)} rows")

# Summary
regions = {}
emails_found = 0
no_website = 0
for c in centers:
    regions[c["region"]] = regions.get(c["region"], 0) + 1
    if c["email"]:
        emails_found += 1
    if not c["website"]:
        no_website += 1

print(f"\nRegion breakdown:")
for r, count in sorted(regions.items()):
    print(f"  {r}: {count}")
print(f"\nEmails found: {emails_found}/{len(centers)}")
print(f"Missing websites: {no_website}")
