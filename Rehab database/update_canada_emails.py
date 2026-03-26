import json
from urllib.parse import urlparse

# Scraped email results from Apify - domain to emails mapping
scraped_data = [
    {"domain":"www.kelburnrecoverycentre.com","emails":["info@kelburnrecoverycentre.com"]},
    {"domain":"addictionrehabtoronto.ca","emails":["info@addictionrehabtoronto.ca"]},
    {"domain":"lastdoor.org","emails":["info@lastdoor.org"]},
    {"domain":"cfhh.ca","emails":["admin@cfhh.ca"]},
    {"domain":"www.bowlinehealth.ca","emails":["info@bowlinehealth.ca"]},
    {"domain":"orchardrecovery.com","emails":["intake@orchardrecovery.com"]},
    {"domain":"andyshouse.com","emails":["info@andyshouse.com"]},
    {"domain":"www.crosbiehousesociety.com","emails":["info@crosbiehouse.ca"]},
    {"domain":"anotherroad.ca","emails":["info@anotherroad.ca"]},
    {"domain":"aurorarecoverycentre.com","emails":["info@aurorarecoverycentre.com"]},
    {"domain":"thefarmrehab.com","emails":["info@thefarmrehab.com"]},
    {"domain":"birchwellnesscenter.ca","emails":["reception@birchwellnesscenter.ca"]},
    {"domain":"renascent.ca","emails":["info@renascent.ca"]},
    {"domain":"stellamarissoberliving.my.canva.site","emails":["stellamarissoberliving@gmail.com"]},
    {"domain":"searidgealcoholrehab.com","emails":["info@searidge.org"]},
    {"domain":"canadiancentreforaddictions.org","emails":["info@ccfatreatment.org"]},
    {"domain":"soberrecovery.ca","emails":["info@heritage-home.com"]},
    {"domain":"ravensview.com","emails":["ravensviewinfo@homewoodhealth.com"]},
    {"domain":"www.thenorthernlights.ca","emails":["info@thenorthernlights.ca"]},
    {"domain":"terradynewellness.ca","emails":["info@terradynewellness.ca"]},
    {"domain":"www.thehealinginstitute.ca","emails":["info@thehealinginstitute.ca"]},
    {"domain":"addictionhealingcentre.ca","emails":["info@addictionhealingcentre.ca"]},
    {"domain":"theresidenceathomewood.com","emails":["residenceinfo@homewoodhealth.com"]},
    {"domain":"www.jellinek.ca","emails":["info@jellinek.ca"]},
    {"domain":"www.sanctuaryvancouverisland.com","emails":["info@sanctuaryvancouverisland.com"]},
    {"domain":"twcrecoverylife.org","emails":["info@twcrecoverylife.org"]},
    {"domain":"greenestone.net","emails":["information@greenestone.net"]},
]

# Build domain -> email map
domain_email = {}
for item in scraped_data:
    d = item["domain"].lower().replace("www.", "")
    emails = [e for e in item["emails"] if "@" in e and "example" not in e and "domain.com" not in e]
    if emails:
        domain_email[d] = emails[0]

# Load centers
with open('canada_centers_data.json', 'r', encoding='utf-8') as f:
    centers = json.load(f)

# Map emails to centers
updated = 0
for c in centers:
    if c['website']:
        parsed = urlparse(c['website'])
        domain = parsed.netloc.lower().replace("www.", "")
        if domain in domain_email:
            c['email'] = domain_email[domain]
            updated += 1

# Save updated JSON
with open('canada_centers_data.json', 'w', encoding='utf-8') as f:
    json.dump(centers, f, indent=2, ensure_ascii=False)

print(f"Updated {updated} centers with emails")
print(f"Centers with emails: {sum(1 for c in centers if c['email'])}")
print(f"Centers without emails: {sum(1 for c in centers if not c['email'])}")

# Now create Excel
try:
    import openpyxl
except ImportError:
    import subprocess
    subprocess.check_call(['pip', 'install', 'openpyxl'])
    import openpyxl

from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side

wb = Workbook()
ws = wb.active
ws.title = "Canada Rehab Centers"

# Headers
headers = ["No.", "Center Name", "Province/Region", "Website", "Phone", "Email", "Recovery.com URL"]
header_font = Font(name='Calibri', bold=True, size=11, color='FFFFFF')
header_fill = PatternFill(start_color='C62828', end_color='C62828', fill_type='solid')
header_alignment = Alignment(horizontal='center', vertical='center', wrap_text=True)
thin_border = Border(
    left=Side(style='thin'),
    right=Side(style='thin'),
    top=Side(style='thin'),
    bottom=Side(style='thin')
)

for col, header in enumerate(headers, 1):
    cell = ws.cell(row=1, column=col, value=header)
    cell.font = header_font
    cell.fill = header_fill
    cell.alignment = header_alignment
    cell.border = thin_border

# Data rows
data_font = Font(name='Calibri', size=10)
link_font = Font(name='Calibri', size=10, color='0563C1', underline='single')
wrap_alignment = Alignment(vertical='center', wrap_text=True)
center_alignment = Alignment(horizontal='center', vertical='center')

# Sort by province then name
centers.sort(key=lambda x: (x['region'], x['centerName']))

for i, c in enumerate(centers, 1):
    row = i + 1
    ws.cell(row=row, column=1, value=i).font = data_font
    ws.cell(row=row, column=1).alignment = center_alignment
    ws.cell(row=row, column=1).border = thin_border

    ws.cell(row=row, column=2, value=c['centerName']).font = data_font
    ws.cell(row=row, column=2).alignment = wrap_alignment
    ws.cell(row=row, column=2).border = thin_border

    ws.cell(row=row, column=3, value=c['region']).font = data_font
    ws.cell(row=row, column=3).alignment = center_alignment
    ws.cell(row=row, column=3).border = thin_border

    cell_web = ws.cell(row=row, column=4, value=c['website'])
    if c['website']:
        cell_web.font = link_font
        cell_web.hyperlink = c['website']
    else:
        cell_web.font = data_font
    cell_web.alignment = wrap_alignment
    cell_web.border = thin_border

    ws.cell(row=row, column=5, value=c['phone']).font = data_font
    ws.cell(row=row, column=5).alignment = center_alignment
    ws.cell(row=row, column=5).border = thin_border

    ws.cell(row=row, column=6, value=c['email']).font = data_font
    ws.cell(row=row, column=6).alignment = wrap_alignment
    ws.cell(row=row, column=6).border = thin_border

    cell_rec = ws.cell(row=row, column=7, value=c['recoveryUrl'])
    cell_rec.font = link_font
    cell_rec.hyperlink = c['recoveryUrl']
    cell_rec.alignment = wrap_alignment
    cell_rec.border = thin_border

    # Alternate row colors
    if i % 2 == 0:
        row_fill = PatternFill(start_color='FFF3F3', end_color='FFF3F3', fill_type='solid')
        for col in range(1, 8):
            ws.cell(row=row, column=col).fill = row_fill

# Column widths
ws.column_dimensions['A'].width = 5
ws.column_dimensions['B'].width = 38
ws.column_dimensions['C'].width = 18
ws.column_dimensions['D'].width = 45
ws.column_dimensions['E'].width = 18
ws.column_dimensions['F'].width = 35
ws.column_dimensions['G'].width = 50

# Freeze top row
ws.freeze_panes = 'A2'

# Auto-filter
ws.auto_filter.ref = f'A1:G{len(centers) + 1}'

wb.save('Canada Rehab Centers.xlsx')
print(f"\nCreated 'Canada Rehab Centers.xlsx' with {len(centers)} centers")
