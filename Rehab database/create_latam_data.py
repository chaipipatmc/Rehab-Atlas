import json
import openpyxl
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from urllib.parse import urlparse

# All 42 centers
centers = [
    {"recoveryUrl":"https://recovery.com/baja-rehab-rosarito-mexico/","centerName":"Baja Rehab","website":"https://bajarehab.com/","phone":"(619)404-4769"},
    {"recoveryUrl":"https://recovery.com/beond-ibogaine-treatment-cancun-mexico/","centerName":"Beond Ibogaine Treatment","website":"https://beondibogaine.com/","phone":"(310)409-7901"},
    {"recoveryUrl":"https://recovery.com/casa-santa-isabel-mexico/","centerName":"Casa Santa Isabel","website":"https://www.casasantaisabel.com/","phone":"(800)706-4901"},
    {"recoveryUrl":"https://recovery.com/centro-izcali-montevideo-uruguay/","centerName":"Centro Izcali","website":"https://izcali.com.uy/","phone":"+59824105479"},
    {"recoveryUrl":"https://recovery.com/baja-ibogaine-rosarito-beach-mexico/","centerName":"Baja Ibogaine Center","website":"https://www.bajaibogainecenter.com/","phone":"(844)660-4600"},
    {"recoveryUrl":"https://recovery.com/clinica-integral-nueva-vida-san-jose-costa-rica/","centerName":"Clinica Integral Nueva Vida","website":"","phone":"(506)2215-6116"},
    {"recoveryUrl":"https://recovery.com/clinica-erimus-parana-floresta-brazil/","centerName":"Clinica Erimus Parana","website":"https://clinicaerimusparana.com.br/","phone":"+55443138-2122"},
    {"recoveryUrl":"https://recovery.com/costa-rica-treatment-center/","centerName":"Costa Rica Treatment Center San Jose","website":"https://costaricatreatmentcenter.com/","phone":"+1(888)981-9092"},
    {"recoveryUrl":"https://recovery.com/costa-rica-new-hope/","centerName":"Costa Rica Recovery","website":"https://costaricarecovery.com/","phone":"+12397468194"},
    {"recoveryUrl":"https://recovery.com/experience-ibogaine-tijuana-mexico/","centerName":"Experience Ibogaine Treatment Center","website":"https://www.experienceibogaine.com/","phone":"+1(866)714-7001"},
    {"recoveryUrl":"https://recovery.com/eira-foundation-tortuguitas-argentina/","centerName":"Eira Foundation","website":"https://fundacioneira.com.ar/","phone":"+54153663-6938"},
    {"recoveryUrl":"https://recovery.com/fundacion-evoluciona-bogota-colombia/","centerName":"Fundacion Evoluciona","website":"https://www.fundacionevoluciona.org/","phone":"(313)3744337"},
    {"recoveryUrl":"https://recovery.com/costa-rica-treatment-riverside-costa-rica/","centerName":"Costa Rica Treatment Center Riverside","website":"https://costaricatreatmentcenter.com/crtc-riverside/","phone":"+1(855)958-5801"},
    {"recoveryUrl":"https://recovery.com/hacienda-del-lago/","centerName":"Hacienda del Lago","website":"https://www.haciendadellago.com.mx/","phone":"(376)766-2220"},
    {"recoveryUrl":"https://recovery.com/holistic-mind-steps-panama-city-panama/","centerName":"Holistic Mind Steps","website":"https://holisticmindsteps.com","phone":"+507831-1446"},
    {"recoveryUrl":"https://recovery.com/iboga-quest-mexico/","centerName":"IbogaQuest","website":"https://www.ibogaquest.com","phone":"(802)748-4600"},
    {"recoveryUrl":"https://recovery.com/genesis-foundation-colombia-bogota-colombia/","centerName":"Genesis Foundation Colombia","website":"https://fundaciongenesis.co/","phone":"+6017026929"},
    {"recoveryUrl":"https://recovery.com/life-oasis-vallarta-puerto-vallarta-mexico/","centerName":"Life Oasis Vallarta","website":"https://lifesoasisvallarta.com/","phone":"+523223495739"},
    {"recoveryUrl":"https://recovery.com/mindscape-retreat/","centerName":"MindScape Retreat","website":"https://www.mindscaperetreat.com/ibogaine-treatment-clinic","phone":"(786)761-8011"},
    {"recoveryUrl":"https://recovery.com/monte-fenix-mexico/","centerName":"Monte Fenix","website":"https://www.montefenix.org.mx/","phone":"+52-55-7100-4456"},
    {"recoveryUrl":"https://recovery.com/las-olas-recovery/","centerName":"Las Olas Recovery","website":"https://www.lasolasrecovery.com/","phone":"+52(612)153-5726"},
    {"recoveryUrl":"https://recovery.com/narconon-colombia/","centerName":"Narconon Colombia","website":"https://www.narcononcolombia.com","phone":"573183309447"},
    {"recoveryUrl":"https://recovery.com/new-life-florianopolis/","centerName":"New Life Florianopolis","website":"https://newlife.rehab/location/luxury-rehab-florianopolis","phone":"(+55)1151785894"},
    {"recoveryUrl":"https://recovery.com/nueva-vida-recovery-home-la-paz-mexico/","centerName":"Nueva Vida Recovery Home","website":"https://nuevavidarecoveryhome.com","phone":"(52)612-102-2611"},
    {"recoveryUrl":"https://recovery.com/oceanica/","centerName":"Oceanica Treatment Center","website":"https://oceanica.com.mx","phone":"+525588547141"},
    {"recoveryUrl":"https://recovery.com/opcion-de-vida-lima-peru/","centerName":"Opcion de Vida","website":"https://opciondevida.org","phone":"+51982568512"},
    {"recoveryUrl":"https://recovery.com/pathways-of-hope-amaguana-ecuador/","centerName":"Pathways of Hope","website":"https://pathways-hope.com/","phone":"(651)278-7113"},
    {"recoveryUrl":"https://recovery.com/pure-life-adventure-tinamaste-costa-rica/","centerName":"Pure Life","website":"https://www.purelifeadventure.com/","phone":"(801)896-9490"},
    {"recoveryUrl":"https://recovery.com/residential/centro-ethos-buenos-aires-argentina/","centerName":"Centro Ethos","website":"","phone":"+544501-7040"},
    {"recoveryUrl":"https://recovery.com/residential/fundacion-kayros-grecia-costa-rica/","centerName":"Fundacion Kayros","website":"","phone":"(+506)8464-3000"},
    {"recoveryUrl":"https://recovery.com/residential/fundacion-manantiales-uruguay-montevideo-uruguay/","centerName":"Fundacion Manantiales Uruguay","website":"","phone":"+59824003400"},
    {"recoveryUrl":"https://recovery.com/residential/fundacion-manantiales-argentina-buenos-aires-argentina/","centerName":"Fundacion Manantiales Argentina","website":"","phone":"+541143828500"},
    {"recoveryUrl":"https://recovery.com/residential/narconon-latin-america/","centerName":"Narconon Latin America","website":"","phone":"+5255-5350-9083"},
    {"recoveryUrl":"https://recovery.com/residential/portal-de-esperanza-lima-peru/","centerName":"Portal de Esperanza","website":"","phone":"+51933822699"},
    {"recoveryUrl":"https://recovery.com/residential/sin-adicciones-peru-lima-peru/","centerName":"Sin Adicciones Peru","website":"","phone":"+511946358146"},
    {"recoveryUrl":"https://recovery.com/residential/sanahra-puerto-vallarta-mexico/","centerName":"Sanahra Residencial Rehab Center","website":"","phone":"+523221936900"},
    {"recoveryUrl":"https://recovery.com/sanctuary-tulum-mexico/","centerName":"Sanctuary Tulum","website":"https://www.sanctuarytulum.com/","phone":"(323)612-2024"},
    {"recoveryUrl":"https://recovery.com/residential/tranquil-hearts-treatment-center/","centerName":"TranquilHearts Treatment Center","website":"","phone":"+52999-747-2604"},
    {"recoveryUrl":"https://recovery.com/the-holistic-sanctuary/","centerName":"The Holistic Sanctuary","website":"https://holisticsanctuary.com","phone":"(844)721-1783"},
    {"recoveryUrl":"https://recovery.com/serenity-vista/","centerName":"Serenity Vista","website":"https://serenityvista.com","phone":"(309)704-4792"},
    {"recoveryUrl":"https://recovery.com/residential/narconon-puebla/","centerName":"Narconon Puebla","website":"","phone":"522223800339"},
    {"recoveryUrl":"https://recovery.com/twilight-recovery-baja-california-mexico/","centerName":"Twilight Recovery Center","website":"https://twilightrecoverycenter.com/","phone":"+1(855)926-0599"},
]

email_map = {
    "bajarehab.com": "team@bajarehab.com",
    "beondibogaine.com": "hello@beond.us",
    "www.casasantaisabel.com": "info@casasantaisabel.com",
    "izcali.com.uy": "contacto@izcali.com.uy",
    "www.bajaibogainecenter.com": "info@bajaibogainecenter.com",
    "costaricatreatmentcenter.com": "info@costaricatreatmentcenter.com",
    "costaricarecovery.com": "info@costarica-recovery.com",
    "www.experienceibogaine.com": "aeden@experienceibogaine.com",
    "www.fundacionevoluciona.org": "info@fundacionevoluciona.org",
    "holisticmindsteps.com": "info@holisticmindsteps.com",
    "www.ibogaquest.com": "info@ibogaquest.com",
    "fundaciongenesis.co": "info@fundaciongenesis.co",
    "lifesoasisvallarta.com": "info@lifesoasisvallarta.com",
    "www.mindscaperetreat.com": "info@mindscaperetreat.com",
    "www.narcononcolombia.com": "info@narconon-colombia.org",
    "nuevavidarecoveryhome.com": "information@nuevavidarecoveryhome.com",
    "opciondevida.org": "info@opciondevida.org",
    "pathways-hope.com": "info@pathways-hope.com",
    "www.purelifeadventure.com": "info@purelifeadventure.com",
    "www.sanctuarytulum.com": "info@sanctuarytulum.com",
    "holisticsanctuary.com": "info@theholisticsanctuary.com",
    "twilightrecoverycenter.com": "info@twilightrecoverycenter.com",
    "newlife.rehab": "info@newlife.rehab",
}

def get_region(url):
    url_lower = url.lower()
    mexico_keywords = ["mexico", "tijuana", "cancun", "tulum", "vallarta", "rosarito", "puebla", "baja", "fenix"]
    if any(k in url_lower for k in mexico_keywords):
        return "Mexico"
    if "costa-rica" in url_lower or "tinamaste" in url_lower:
        return "Costa Rica"
    if "panama" in url_lower:
        return "Panama"
    if "colombia" in url_lower or "bogota" in url_lower:
        return "Colombia"
    if "brazil" in url_lower or "florianopolis" in url_lower or "parana" in url_lower:
        return "Brazil"
    if "argentina" in url_lower or "buenos-aires" in url_lower or "tortuguitas" in url_lower:
        return "Argentina"
    if "peru" in url_lower or "lima" in url_lower:
        return "Peru"
    if "ecuador" in url_lower or "amaguana" in url_lower:
        return "Ecuador"
    if "uruguay" in url_lower or "montevideo" in url_lower:
        return "Uruguay"
    return "Latin America"

def get_email(website):
    if not website:
        return ""
    parsed = urlparse(website)
    domain = parsed.netloc or parsed.path
    domain = domain.rstrip("/")
    if domain in email_map:
        return email_map[domain]
    if domain.startswith("www."):
        bare = domain[4:]
        if bare in email_map:
            return email_map[bare]
    www = "www." + domain
    if www in email_map:
        return email_map[www]
    parts = domain.split(".")
    if len(parts) > 2:
        base = ".".join(parts[-2:])
        if base in email_map:
            return email_map[base]
    return ""

# Build enriched data
for c in centers:
    c["region"] = get_region(c["recoveryUrl"])
    c["email"] = get_email(c["website"])

# Sort by region then center name
centers.sort(key=lambda x: (x["region"], x["centerName"].lower()))

# Save JSON
with open("latam_centers_data.json", "w", encoding="utf-8") as f:
    json.dump(centers, f, indent=2, ensure_ascii=False)
print(f"Saved latam_centers_data.json with {len(centers)} centers")

# Create Excel
wb = openpyxl.Workbook()
ws = wb.active
ws.title = "Latin America Rehab Centers"

headers = ["#", "Region", "Center Name", "Website", "Phone", "Recovery.com URL", "Email"]
header_font = Font(name="Calibri", bold=True, color="FFFFFF", size=11)
header_fill = PatternFill(start_color="2F5496", end_color="2F5496", fill_type="solid")
header_alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)

thin_border = Border(
    left=Side(style="thin", color="B4C6E7"),
    right=Side(style="thin", color="B4C6E7"),
    top=Side(style="thin", color="B4C6E7"),
    bottom=Side(style="thin", color="B4C6E7"),
)

# Write headers
for col_idx, header in enumerate(headers, 1):
    cell = ws.cell(row=1, column=col_idx, value=header)
    cell.font = header_font
    cell.fill = header_fill
    cell.alignment = header_alignment
    cell.border = thin_border

# Alternating row fills
light_fill = PatternFill(start_color="D6E4F0", end_color="D6E4F0", fill_type="solid")
white_fill = PatternFill(start_color="FFFFFF", end_color="FFFFFF", fill_type="solid")
yellow_fill = PatternFill(start_color="FFFF00", end_color="FFFF00", fill_type="solid")

# Write data
for idx, c in enumerate(centers):
    row = idx + 2
    row_fill = light_fill if idx % 2 == 0 else white_fill

    ws.cell(row=row, column=1, value=idx + 1).font = Font(name="Calibri", size=11)
    ws.cell(row=row, column=1).alignment = Alignment(horizontal="center")
    ws.cell(row=row, column=2, value=c["region"]).font = Font(name="Calibri", size=11)
    ws.cell(row=row, column=3, value=c["centerName"]).font = Font(name="Calibri", size=11)

    # Website - yellow if missing
    web_cell = ws.cell(row=row, column=4, value=c["website"])
    web_cell.font = Font(name="Calibri", size=11)
    if not c["website"]:
        web_cell.fill = yellow_fill

    ws.cell(row=row, column=5, value=c["phone"]).font = Font(name="Calibri", size=11)
    ws.cell(row=row, column=6, value=c["recoveryUrl"]).font = Font(name="Calibri", size=11)
    ws.cell(row=row, column=7, value=c["email"]).font = Font(name="Calibri", size=11)

    # Apply row fill and border
    for col in range(1, 8):
        cell = ws.cell(row=row, column=col)
        if col != 4 or c["website"]:
            cell.fill = row_fill
        elif not c["website"]:
            cell.fill = yellow_fill
        cell.border = thin_border

# Column widths
col_widths = {"A": 5, "B": 14, "C": 42, "D": 55, "E": 22, "F": 65, "G": 32}
for col_letter, width in col_widths.items():
    ws.column_dimensions[col_letter].width = width

# Auto-filter
ws.auto_filter.ref = f"A1:G{len(centers) + 1}"

# Freeze panes
ws.freeze_panes = "A2"

wb.save("Latin America Rehab Centers.xlsx")
print(f"Saved Latin America Rehab Centers.xlsx with {len(centers)} rows")

# Summary
regions = {}
emails_found = 0
no_website = 0
for c in centers:
    regions[c["region"]] = regions.get(c["region"], 0) + 1
    if c["email"]:
        emails_found += 1
    if not c["website"]:
        no_website += 1

print(f"\nRegion breakdown:")
for r, count in sorted(regions.items()):
    print(f"  {r}: {count}")
print(f"\nEmails found: {emails_found}/{len(centers)}")
print(f"Missing websites: {no_website}")
